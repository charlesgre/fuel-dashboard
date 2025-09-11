# arbs_tab.py
from __future__ import annotations
import io, re
from dataclasses import dataclass
from typing import Dict, Tuple, Optional
from datetime import datetime, date, time
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

DEFAULT_XLSX_PATH = "Arbs/Arbsheet updated.xlsx"
DEFAULT_SHEET = "MAIN"

# just under the imports / constants
SIGN_COLOR_ROW_LABELS = {
    "rmg", "rmg med", "rmg us",
    "0.5%", "0.5 med", "0.5 us",
    "0.5", "0.50"   # <- important
}

@dataclass(frozen=True)
class Merge:
    min_col: int
    min_row: int
    max_col: int
    max_row: int

# ---------- helpers couleurs ----------
def _hex_color(c) -> str | None:
    if not c:
        return None
    try:
        rgb = c.rgb
        if not rgb:
            return None
        if len(rgb) == 8:  # AARRGGBB
            return f"#{rgb[2:]}"
        if len(rgb) == 6:
            return f"#{rgb}"
    except Exception:
        pass
    return None

def _is_dark(hex_color: str) -> bool:
    if not hex_color or not hex_color.startswith("#") or len(hex_color) != 7:
        return False
    r = int(hex_color[1:3], 16)
    g = int(hex_color[3:5], 16)
    b = int(hex_color[5:7], 16)
    lum = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return lum < 80

def _border_css(b) -> str:
    if not b:
        return ""
    parts = []
    for side_name in ("left", "right", "top", "bottom"):
        side = getattr(b, side_name, None)
        if side and side.style:
            width = "1px"; style = "solid"
            if side.style in ("medium", "thick"): width = "2px"
            if side.style in ("hair", "dotted", "dashDot"): style = "dotted"
            color = _hex_color(getattr(side, "color", None)) or "#BEBEBE"
            parts.append(f"border-{side_name}:{width} {style} {color};")
    return "".join(parts)

def _align_css(al):
    if not al:
        return ""
    parts = []
    if al.horizontal: parts.append(f"text-align:{al.horizontal};")
    if al.vertical:   parts.append(f"vertical-align:{ {'center':'middle'}.get(al.vertical, al.vertical) };")
    parts.append("white-space:normal;" if getattr(al, "wrap_text", False) else "white-space:nowrap;")
    return "".join(parts)

def _font_css(f, force_white: bool=False):
    parts = []
    if f and f.bold:   parts.append("font-weight:700;")
    if f and f.italic: parts.append("font-style:italic;")
    if f and f.size:   parts.append(f"font-size:{f.size}px;")
    if force_white:
        parts.append("color:#ffffff;")
    else:
        col = _hex_color(getattr(f, "color", None))
        if col: parts.append(f"color:{col};")
    if not parts: parts.append("font-size:13px;")
    return "".join(parts)

def _fill_css(fill, fallback_blue=False):
    try:
        col = _hex_color(fill.fgColor) if fill and fill.fgColor else None
        if fallback_blue:  # titres/fonds sombres → bleu lisible
            return "background-color:#2c5282;"
        if col:
            return f"background-color:{col};"
    except Exception:
        pass
    return ""

# ---------- formatage valeurs ----------
def _decimals_from_number_format(fmt: str | None) -> Optional[int]:
    """Essaye d'inférer le nombre de décimales d'après le format Excel."""
    if not fmt:
        return None
    # exemples: "0", "0.0", "0.00", "#,##0.000", "0.00_);[Red](0.00)"
    m = re.search(r"\.(0+)", fmt)
    if m:
        return len(m.group(1))
    return None

def _format_cell_value(cell) -> str:
    """Formate la valeur comme Excel dans tes tableaux: dates dd.MMM, chiffres avec décimales correctes."""
    v = cell.value
    if v is None:
        return ""
    # dates
    if isinstance(v, (datetime, date)):
        d: datetime = v if isinstance(v, datetime) else datetime.combine(v, time.min)
        return d.strftime("%d.%b")
    # texte direct
    if isinstance(v, str):
        # tente parsing de date ISO provenant d'Excel (ex: "2025-09-30 00:00:00")
        try:
            d = datetime.fromisoformat(v.strip())
            return d.strftime("%d.%b")
        except Exception:
            return v.replace("\n", "<br>")
    # nombres
    if isinstance(v, (int, float)):
        dec = _decimals_from_number_format(getattr(cell, "number_format", None))
        if dec is None:
            # par défaut: int → 0 décimales, float → 2 décimales
            dec = 0 if isinstance(v, int) or float(v).is_integer() else 2
        if dec == 0:
            return f"{int(round(v)):,}".replace(",", " ")
        return f"{v:,.{dec}f}".replace(",", " ")
    # fallback
    return str(v).replace("\n", "<br>")

def _value_sign_bg(cell, force: bool = False) -> str:
    """
    Renvoie un background rouge/vert selon le signe.
    - Si force=False : ne touche pas aux cellules qui ont déjà un fill Excel.
    - Si force=True  : applique le rouge/vert même s'il y a déjà un fill (écrase).
    """
    v = cell.value
    if not force:
        has_explicit_bg = bool(_hex_color(getattr(getattr(cell, "fill", None), "fgColor", None)))
        if has_explicit_bg:
            return ""
    if isinstance(v, (int, float)):
        if v < 0:
            return "background-color:#fca5a5; color:#000000;"  # rouge
        if v > 0:
            return "background-color:#86efac; color:#000000;"  # vert
    return ""

# ---------- helpers sheet ----------
def _is_col_hidden(ws, idx: int) -> bool:
    cd = ws.column_dimensions.get(get_column_letter(idx))
    return bool(cd and cd.hidden)

def _is_row_hidden(ws, idx: int) -> bool:
    rd = ws.row_dimensions.get(idx)
    return bool(rd and rd.hidden)

def _is_empty(v) -> bool:
    return v is None or (isinstance(v, str) and v.strip() == "")

def _detect_right_cut(ws, min_r, max_r, min_c, max_c,
                      look_rows: int = 120, empty_run_needed: int = 5) -> int:
    look_to_row = min(max_r, min_r + look_rows - 1)
    last_val_col = min_c
    empty_run = 0
    for c in range(min_c, max_c + 1):
        if _is_col_hidden(ws, c):
            continue
        has_val = False
        for r in range(min_r, look_to_row + 1):
            if _is_row_hidden(ws, r):
                continue
            if not _is_empty(ws.cell(r, c).value):
                has_val = True
                break
        if has_val:
            last_val_col = c
            empty_run = 0
        else:
            empty_run += 1
            if empty_run >= empty_run_needed:
                break
    return last_val_col

def _visible_bounds(ws, col_cap: Optional[int] = None):
    max_r = ws.max_row
    max_c = ws.max_column
    min_r = 1
    while min_r <= max_r and _is_row_hidden(ws, min_r):
        min_r += 1
    while max_r >= 1 and _is_row_hidden(ws, max_r):
        max_r -= 1
    min_c = 1
    while min_c <= max_c and _is_col_hidden(ws, min_c):
        min_c += 1
    detected_end = _detect_right_cut(
        ws, min_r, max_r, min_c, max_c, look_rows=120, empty_run_needed=5
    )
    max_c = detected_end
    if col_cap is None:
        max_c = min(max_c, 18)
    elif isinstance(col_cap, int) and col_cap > 0:
        max_c = min(max_c, min_c + col_cap - 1)
    return max(min_r, 1), max_r, max(min_c, 1), max_c

def _collect_merges(ws) -> Dict[Tuple[int, int], Merge]:
    merges: Dict[Tuple[int,int], Merge] = {}
    for rng in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = range_boundaries(str(rng))
        if any(_is_row_hidden(ws, r) for r in range(min_r, max_r+1)):  continue
        if any(_is_col_hidden(ws, c) for c in range(min_c, max_c+1)):  continue
        merges[(min_r, min_c)] = Merge(min_c, min_r, max_c, max_r)
    return merges

def _extract_dataframe(ws, col_cap: Optional[int]) -> pd.DataFrame:
    min_r, max_r, min_c, max_c = _visible_bounds(ws, col_cap)
    data = []
    for r in range(min_r, max_r+1):
        if _is_row_hidden(ws, r): continue
        row_vals = []
        for c in range(min_c, max_c+1):
            if _is_col_hidden(ws, c): continue
            row_vals.append(ws.cell(r, c).value)
        data.append(row_vals)
    cols = [get_column_letter(c) for c in range(min_c, max_c+1) if not _is_col_hidden(ws, c)]
    return pd.DataFrame(data, columns=cols)

def _sheet_to_html(ws, col_cap: Optional[int]) -> str:
    min_r, max_r, min_c, max_c = _visible_bounds(ws, col_cap)
    merges = _collect_merges(ws)
    def _merge_covering(r: int, c: int) -> Optional[Merge]:
        for (tl_r, tl_c), m in merges.items():
            if m.min_row <= r <= m.max_row and m.min_col <= c <= m.max_col:
                return m
        return None
    def _find_label_col(r: int) -> int:
        scan_to = min(max_c, min_c + 4)
        for c in range(min_c, scan_to + 1):
            if _is_col_hidden(ws, c): continue
            cell = ws.cell(r, c)
            m = _merge_covering(r, c)
            if m:
                anchor_val = ws.cell(m.min_row, m.min_col).value
                if not _is_empty(anchor_val):
                    return c
            if not _is_empty(cell.value):
                return c
        return min_c
    html = [
        '<div style="overflow:auto; max-height:80vh; border:1px solid #ddd; padding:8px;">',
        """
        <style>
        .arbs-table { border-collapse:collapse; font-family:Inter,system-ui,Arial; font-size:13px; }
        .arbs-table tr:nth-child(even) td { background-color:#fafafa; }
        .arbs-table td { padding:4px 8px; }
        </style>
        """,
        '<table class="arbs-table">'
    ]
    covered = set()
    for (tl_r, tl_c), m in merges.items():
        for rr in range(m.min_row, m.max_row + 1):
            for cc in range(m.min_col, m.max_col + 1):
                if not (rr == tl_r and cc == tl_c):
                    covered.add((rr, cc))
    for r in range(min_r, max_r + 1):
        if _is_row_hidden(ws, r): continue
        html.append("<tr>")
        label_col = _find_label_col(r)
        m_first = _merge_covering(r, label_col)
        if m_first:
            label_cell = ws.cell(m_first.min_row, m_first.min_col)
            row_label = re.sub(r"\s+", " ", str(label_cell.value or "").strip().lower())
            on_anchor_row = (r == m_first.min_row)
        else:
            label_cell = ws.cell(r, label_col)
            row_label = re.sub(r"\s+", " ", str(label_cell.value or "").strip().lower())
            on_anchor_row = True
        for c in range(min_c, max_c + 1):
            if _is_col_hidden(ws, c) or (r, c) in covered: continue
            cell = ws.cell(r, c)
            m = merges.get((r, c))
            rs = (m.max_row - m.min_row + 1) if m else 1
            cs = (m.max_col - m.min_col + 1) if m else 1
            raw_bg = _hex_color(getattr(getattr(cell, "fill", None), "fgColor", None))
            dark_bg = _is_dark(raw_bg) if raw_bg else False
            use_blue = bool(getattr(cell.font, "bold", False)) or dark_bg
            styles = []
            styles.append(_border_css(cell.border))
            styles.append(_align_css(cell.alignment))
            styles.append(_fill_css(cell.fill, fallback_blue=use_blue))
            allow_sign_color = (on_anchor_row and row_label in SIGN_COLOR_ROW_LABELS and c != label_col)
            if allow_sign_color:
                styles.append(_value_sign_bg(cell, force=True))
            styles.append(_font_css(cell.font, force_white=(use_blue or dark_bg)))
            if not cell.border or not any(getattr(cell.border, s).style for s in ("left","right","top","bottom")):
                styles.append("border:1px solid #e6e6e6;")
            value_html = _format_cell_value(cell)
            html.append(f'<td rowspan="{rs}" colspan="{cs}" style="{"".join(styles)}">{value_html}</td>')
        html.append("</tr>")
    html.append("</table></div>")
    return "".join(html)

# ===================== STREAMLIT TAB =====================
def render():
    st.header("Arbs")
    path = st.text_input("Chemin du fichier Excel des arbs :", DEFAULT_XLSX_PATH)
    col1, col2, col3 = st.columns([1,1,1], gap="small")
    editable   = col1.toggle("Mode éditable (grid)", value=False)
    show_export= col2.toggle("Activer export XLSX", value=True)
    col_cap = col3.number_input(
        "Max colonnes à afficher (18 = par défaut)",
        min_value=0, max_value=200, value=18, step=1
    )
    cap = int(col_cap) if col_cap and col_cap > 0 else None
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        st.error(f"Impossible d’ouvrir le fichier : {e}")
        st.stop()
    visible_sheets = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        st.warning("Aucune feuille visible dans ce classeur.")
        st.stop()
    sheet = st.selectbox(
        "Feuille source :", visible_sheets,
        index=visible_sheets.index(DEFAULT_SHEET) if DEFAULT_SHEET in visible_sheets else 0
    )
    ws = wb[sheet]
    cap = int(col_cap) if col_cap and col_cap > 0 else None
    if editable:
        st.caption("Vue éditable (valeurs uniquement).")
        df = _extract_dataframe(ws, cap)
        edited = st.data_editor(df, use_container_width=True, height=650)
        if show_export:
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                edited.to_excel(writer, sheet_name="Arbs", index=False)
            st.download_button(
                "Exporter la vue éditée en XLSX", data=out.getvalue(),
                file_name="Arbs_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.caption("Rendu fidèle d’Excel (fusions, styles, couleurs plus visibles + règles Excel).")
        html = _sheet_to_html(ws, cap)
        st.markdown(html, unsafe_allow_html=True)
        if show_export:
            df = _extract_dataframe(ws, cap)
            out = io.BytesIO()
            with pd.ExcelWriter(out, engine="openpyxl") as writer:
                df.to_excel(writer, sheet_name="Arbs_visible", index=False)
            st.download_button(
                "Exporter la zone visible en XLSX", data=out.getvalue(),
                file_name="Arbs_visible.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

if __name__ == "__main__":
    import streamlit.web.bootstrap
    def main():
        render()
    streamlit.web.bootstrap.run(main, "", [], flag_options={})
